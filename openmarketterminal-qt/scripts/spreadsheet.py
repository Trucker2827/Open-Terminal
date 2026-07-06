"""
spreadsheet.py — unified local + Google Sheets I/O for the OpenMarketTerminal workflow node.

Usage:
    python spreadsheet.py --args '<json>'

The JSON args object supports:
  source           : "local" | "google"
  operation        : "read" | "write" | "append" | "clear"
  file_path        : path to local .xlsx or .csv  (source=local)
  spreadsheet_id   : Google Sheet ID from URL     (source=google)
  sheet_name       : sheet/tab name (default "Sheet1")
  range            : cell range e.g. "A1:D100"    (default "A1:Z1000")
  credentials_path : path to service account credentials.json
                     leave empty for public sheets (source=google, read-only)
  data             : [[...], [...], ...]  rows to write/append

Outputs a single JSON object to stdout.
"""

import sys
import json
import argparse
import os
import datetime
import csv


def parse_range(range_str):
    """Parse 'A1:D100' -> (r1, c1, r2, c2) 1-based. Returns None on failure."""
    try:
        parts = range_str.upper().split(":")
        if len(parts) != 2:
            return None

        def parse_cell(cell):
            col = 0
            i = 0
            while i < len(cell) and cell[i].isalpha():
                col = col * 26 + (ord(cell[i]) - ord("A") + 1)
                i += 1
            row = int(cell[i:])
            return row, col

        r1, c1 = parse_cell(parts[0].strip())
        r2, c2 = parse_cell(parts[1].strip())
        return r1, c1, r2, c2
    except Exception:
        return None


# ── LOCAL ─────────────────────────────────────────────────────────────────────

def json_safe_value(value):
    """Convert openpyxl values into JSON-safe scalars while preserving formulas."""
    if value is None:
        return None
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return value.isoformat()
    return value


def local_read(file_path, sheet_name, range_str):
    ext = os.path.splitext(file_path)[1].lower()
    parsed = parse_range(range_str)
    r1, c1, r2, c2 = parsed if parsed else (1, 1, 1000, 26)

    if ext == ".csv":
        rows = []
        with open(file_path, "r", encoding="utf-8", newline="") as f:
            reader = csv.reader(f)
            for line_no, cells in enumerate(reader, start=1):
                if line_no < r1:
                    continue
                if line_no > r2:
                    break
                rows.append(cells[c1 - 1 : c2])
        return {"source": "local", "operation": "read", "path": file_path,
                "rows": len(rows), "data": rows}

    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    rows = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=r1, max_row=r2,
                                              min_col=c1, max_col=c2,
                                              values_only=True), start=r1):
        if any(v is not None for v in row):
            rows.append([v if v is not None else None for v in row])
    wb.close()
    return {"source": "local", "operation": "read", "path": file_path,
            "rows": len(rows), "data": rows}


def local_read_workbook(file_path):
    """Read every worksheet, preserving formulas and trimming blank trailing rows/cols."""
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=False, data_only=False)
    sheets = []
    for ws in wb.worksheets:
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1
        rows = []
        last_row = -1
        last_col = -1
        for r in range(1, max_row + 1):
            row_values = []
            row_has_data = False
            for c in range(1, max_col + 1):
                value = json_safe_value(ws.cell(row=r, column=c).value)
                if value is not None and value != "":
                    row_has_data = True
                    last_col = max(last_col, c - 1)
                row_values.append(value)
            if row_has_data:
                last_row = r - 1
            rows.append(row_values)

        if last_row >= 0 and last_col >= 0:
            rows = [row[: last_col + 1] for row in rows[: last_row + 1]]
        else:
            rows = []
        sheets.append({"name": ws.title, "rows": len(rows), "data": rows})
    wb.close()
    return {"source": "local", "operation": "read_workbook", "path": file_path,
            "sheet_count": len(sheets), "sheets": sheets}


def local_write(file_path, sheet_name, range_str, data, append=False):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        mode = "a" if append and os.path.exists(file_path) else "w"
        with open(file_path, mode, encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerows(data)
        return {"source": "local", "operation": "append" if append else "write",
                "path": file_path, "rows_written": len(data)}

    import openpyxl
    parsed = parse_range(range_str)
    r1, c1 = (parsed[0], parsed[1]) if parsed else (1, 1)

    if append and os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
    else:
        wb = openpyxl.Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        # Remove default empty sheet if it exists
        if "Sheet" in wb.sheetnames and ws.title != "Sheet":
            del wb["Sheet"]

    start_row = r1
    if append:
        start_row = ws.max_row + 1 if ws.max_row else 1

    rows_written = 0
    for ri, row in enumerate(data):
        for ci, val in enumerate(row):
            ws.cell(row=start_row + ri, column=c1 + ci, value=val)
        rows_written += 1

    wb.save(file_path)
    return {"source": "local", "operation": "append" if append else "write",
            "path": file_path, "rows_written": rows_written}


def safe_sheet_title(name, existing):
    title = str(name or "Sheet").strip()[:31] or "Sheet"
    for ch in "[]:*?/\\":
        title = title.replace(ch, "_")
    base = title
    suffix = 2
    while title in existing:
        tail = f"_{suffix}"
        title = (base[: 31 - len(tail)] + tail) if len(base) + len(tail) > 31 else base + tail
        suffix += 1
    existing.add(title)
    return title


def local_write_workbook(file_path, sheets):
    import openpyxl
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)
    existing = set()
    sheet_count = 0
    rows_written = 0

    for sheet in sheets:
        if not isinstance(sheet, dict):
            continue
        title = safe_sheet_title(sheet.get("name", f"Sheet{sheet_count + 1}"), existing)
        ws = wb.create_sheet(title)
        data = sheet.get("data", [])
        if not isinstance(data, list):
            data = []
        for ri, row in enumerate(data, start=1):
            if not isinstance(row, list):
                row = [row]
            for ci, value in enumerate(row, start=1):
                if value is None or value == "":
                    continue
                ws.cell(row=ri, column=ci, value=value)
            rows_written += 1
        sheet_count += 1

    if sheet_count == 0:
        wb.create_sheet("Sheet1")
        sheet_count = 1

    wb.save(file_path)
    return {"source": "local", "operation": "write_workbook",
            "path": file_path, "sheet_count": sheet_count, "rows_written": rows_written}


def local_clear(file_path, sheet_name, range_str):
    import openpyxl
    parsed = parse_range(range_str)
    r1, c1, r2, c2 = parsed if parsed else (1, 1, 1000, 26)

    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).value = None
    wb.save(file_path)
    return {"source": "local", "operation": "clear", "path": file_path}


# ── GOOGLE SHEETS ─────────────────────────────────────────────────────────────

def _get_gspread_client(credentials_path):
    import gspread
    if credentials_path:
        return gspread.service_account(filename=credentials_path)
    else:
        # Anonymous / public read-only via API key not supported by gspread.
        # Fall back to requests for public sheets.
        return None


def google_read(spreadsheet_id, sheet_name, range_str, credentials_path):
    import gspread
    if credentials_path:
        gc = gspread.service_account(filename=credentials_path)
        sh = gc.open_by_key(spreadsheet_id)
        ws = sh.worksheet(sheet_name)
        full_range = f"{sheet_name}!{range_str}"
        rows = ws.get(range_str)
    else:
        # Public sheet — use requests directly (no auth needed)
        import requests
        url = (f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}"
               f"/values/{sheet_name}!{range_str}")
        resp = requests.get(url)
        if not resp.ok:
            return {"error": f"HTTP {resp.status_code}: {resp.text[:200]}"}
        rows = resp.json().get("values", [])

    return {"source": "google", "operation": "read", "sheet_id": spreadsheet_id,
            "range": range_str, "rows": len(rows), "data": rows}


def google_write(spreadsheet_id, sheet_name, range_str, data, credentials_path, append=False):
    import gspread
    gc = gspread.service_account(filename=credentials_path)
    sh = gc.open_by_key(spreadsheet_id)
    ws = sh.worksheet(sheet_name)
    if append:
        result = ws.append_rows(data, value_input_option="USER_ENTERED",
                                insert_data_option="INSERT_ROWS")
        updated = result.get("updates", {}).get("updatedCells", len(data))
    else:
        result = ws.update(range_str, data, value_input_option="USER_ENTERED")
        updated = result.get("updatedCells", len(data))
    return {"source": "google", "operation": "append" if append else "write",
            "sheet_id": spreadsheet_id, "updated_cells": updated}


def google_clear(spreadsheet_id, sheet_name, range_str, credentials_path):
    import gspread
    gc = gspread.service_account(filename=credentials_path)
    sh = gc.open_by_key(spreadsheet_id)
    ws = sh.worksheet(sheet_name)
    ws.batch_clear([range_str])
    return {"source": "google", "operation": "clear", "sheet_id": spreadsheet_id}


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--args", required=True)
    parsed = parser.parse_args()

    try:
        args = json.loads(parsed.args)
    except json.JSONDecodeError as e:
        print(json.dumps({"error": f"Invalid --args JSON: {e}"}))
        sys.exit(1)

    source           = args.get("source", "local")
    operation        = args.get("operation", "read")
    file_path        = args.get("file_path", "")
    spreadsheet_id   = args.get("spreadsheet_id", "")
    sheet_name       = args.get("sheet_name", "Sheet1")
    range_str        = args.get("range", "A1:Z1000")
    credentials_path = args.get("credentials_path", "")
    data             = args.get("data", [])
    sheets           = args.get("sheets", [])

    # Normalise data: may be a list of rows or a wrapped object
    if isinstance(data, dict) and "data" in data:
        data = data["data"]
    if not isinstance(data, list):
        data = []

    try:
        if source == "local":
            if not file_path:
                result = {"error": "file_path is required for source=local"}
            elif operation == "read_workbook":
                result = local_read_workbook(file_path)
            elif operation == "write_workbook":
                result = local_write_workbook(file_path, sheets)
            elif operation == "read":
                result = local_read(file_path, sheet_name, range_str)
            elif operation in ("write", "append"):
                result = local_write(file_path, sheet_name, range_str, data,
                                     append=(operation == "append"))
            elif operation == "clear":
                result = local_clear(file_path, sheet_name, range_str)
            else:
                result = {"error": f"Unknown operation: {operation}"}

        elif source == "google":
            if not spreadsheet_id:
                result = {"error": "spreadsheet_id is required for source=google"}
            elif operation == "read":
                result = google_read(spreadsheet_id, sheet_name, range_str, credentials_path)
            elif operation in ("write", "append"):
                if not credentials_path:
                    result = {"error": "credentials_path is required for write/append to Google Sheets"}
                else:
                    result = google_write(spreadsheet_id, sheet_name, range_str, data,
                                          credentials_path, append=(operation == "append"))
            elif operation == "clear":
                if not credentials_path:
                    result = {"error": "credentials_path is required for clear on Google Sheets"}
                else:
                    result = google_clear(spreadsheet_id, sheet_name, range_str, credentials_path)
            else:
                result = {"error": f"Unknown operation: {operation}"}
        else:
            result = {"error": f"Unknown source: {source}"}

    except Exception as e:
        result = {"error": str(e)}

    print(json.dumps(result))


if __name__ == "__main__":
    main()
